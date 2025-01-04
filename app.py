import json
import requests
import base64
import urllib
import os
import openpyxl

API_KEY = ""
SECRET_KEY = ""
FILE_NAME = "发票数据.xlsx"

def get_all_file():
    file_list = []
    for root, _, files in os.walk("."):
        for file in files:
            if file.lower().endswith(('.jpg', '.png', '.pdf')):
                relative_path = os.path.relpath(os.path.join(root, file))
                file_list.append(relative_path)
    return file_list

def rb(path):
    with open(path,'rb')as f:
        return urllib.parse.quote_plus(base64.b64encode(f.read()).decode("utf-8"))

def get_data(file):
    url = "https://aip.baidubce.com/rest/2.0/ocr/v1/multiple_invoice?access_token=" + get_access_token()
    _, ext = os.path.splitext(file)
    file_type=ext.lower().replace('.','')
    if file_type=='pdf':
        way='pdf_file'
    else:
        way='image'
    payload = '%s=%s&verify_parameter=false&probability=false&location=false' %(way,rb(file))
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Accept': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=payload.encode("utf-8"))
    return json.loads(response.text)

def get_access_token():
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {"grant_type": "client_credentials", "client_id": API_KEY, "client_secret": SECRET_KEY}
    return str(requests.post(url, params=params).json().get("access_token"))

def into_xlsx(file):
    def safe_get(data, key, default=""):
        return data.get(key, [{}])[0].get("word", default) if data.get(key) else default
    result = file.get("words_result", [])[0].get("result", {})
    data = {
        "发票号码": safe_get(result, "InvoiceNum"),
        "发票代码": safe_get(result, "InvoiceCode"),
        "开票日期": safe_get(result, "InvoiceDate"),
        "发票类型": safe_get(result, "InvoiceType"),
        "购买方名称": safe_get(result, "PurchaserName"),
        "购买方税号": safe_get(result, "PurchaserRegisterNum"),
        "销售方名称": safe_get(result, "SellerName"),
        "销售方税号": safe_get(result, "SellerRegisterNum"),
        "金额大写": safe_get(result, "AmountInWords"),
        "金额小写": safe_get(result, "AmountInFiguers"),
        "不含税金额": safe_get(result, "TotalAmount"),
        "税额": safe_get(result, "TotalTax"),
        "商品名称": safe_get(result, "CommodityName"),
        "商品数量": safe_get(result, "CommodityNum"),
        "商品单价": safe_get(result, "CommodityPrice"),
        "商品金额": safe_get(result, "CommodityAmount"),
        "商品税率": safe_get(result, "CommodityTaxRate"),
        "商品税额": safe_get(result, "CommodityTax"),
        "开票人": safe_get(result, "NoteDrawer"),
    }
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active
    start_row = sheet.max_row + 1
    for col, value in enumerate(data.values(), start=1):
        sheet.cell(row=start_row, column=col, value=value)
    workbook.save(FILE_NAME)

def main():
    for i in get_all_file():
        into_xlsx(get_data(i))


main()